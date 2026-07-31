"""Tracking spreadsheet + visual dashboard, generated from the ledger.

Sharper's operations discipline, as tooling: every decision the engine makes
(tickets, limit orders, passes, refusals, flags) and every graded fair line
lives in data/track_record.json; this module renders that single source of
truth into

- ``data/tracker.xlsx`` — a formula-driven workbook (Bets, FairLines,
  Dashboard) that recomputes win rate, ROI, CLV, and calibration stats when
  rows are added or settled; and
- a self-contained HTML dashboard for the published tracker artifact.

Regenerate both with ``sharpods-tracker`` after each settle; the numbers can
never drift from the ledger because they are derived from it on every run.
"""

from __future__ import annotations

import argparse
import base64
import json
from importlib import resources
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from sharpods.ledger import load_record

ARIAL = "Arial"
BLUE = "FF0000FF"  # hardcoded inputs
BLACK = "FF000000"  # formulas
YELLOW = "FFFFFF00"  # cells the user may edit


# ---------------------------------------------------------------------------
# Dataset
# ---------------------------------------------------------------------------


def dataset(record: dict[str, Any]) -> dict[str, Any]:
    """Flatten the ledger into decisions, fair-line grades, and summary
    stats for rendering. The xlsx recomputes the same stats with formulas;
    this Python copy feeds the HTML dashboard."""
    decisions: list[dict[str, Any]] = []
    fairlines: list[dict[str, Any]] = []
    for slate in record.get("slates", []):
        for d in slate.get("decisions", []):
            decisions.append({"date": slate["date"], **d})
        for g in slate.get("games", []):
            # Gradeable = we predicted AND the close and result are known.
            # Refused markets carry a close but no prediction; pending games
            # carry a prediction but no result. Both are skipped, not guessed.
            if (
                g.get("our_home_fair") is None
                or g.get("close_home_novig") is None
                or g.get("home_won") is None
            ):
                continue
            fairlines.append(
                {
                    "date": slate["date"],
                    "event": g["event_id"],
                    "our_fair": g["our_home_fair"],
                    "close_novig": g["close_home_novig"],
                    "err_pts": round(
                        abs(g["our_home_fair"] - g["close_home_novig"]) * 100, 2
                    ),
                    "home_won": bool(g["home_won"]),
                    "flag": g.get("source_flag", "clean"),
                    "final": g.get("final", ""),
                }
            )

    tickets = [d for d in decisions if d["type"] == "ticket"]
    settled_tickets = [t for t in tickets if t["result"] in ("won", "lost")]
    wins = sum(1 for t in settled_tickets if t["result"] == "won")
    staked = sum(float(t.get("stake") or 0) for t in settled_tickets)
    pnl = sum(float(t.get("pnl") or 0) for t in settled_tickets)
    clv_vals = [t["clv_novig"] for t in settled_tickets if t.get("clv_novig") is not None]

    settled = [d for d in decisions if d["result"] not in ("pending",)]
    worked = [d for d in settled if d["process"] == "worked"]
    missed = [d for d in settled if d["process"] == "miss"]

    clean = [f["err_pts"] for f in fairlines if f["flag"] == "clean"]
    flagged = [f["err_pts"] for f in fairlines if f["flag"] != "clean"]
    briers_ours = [
        (f["our_fair"] - (1.0 if f["home_won"] else 0.0)) ** 2 for f in fairlines
    ]
    briers_close = [
        (f["close_novig"] - (1.0 if f["home_won"] else 0.0)) ** 2 for f in fairlines
    ]

    def avg(xs: list[float]) -> float | None:
        return round(sum(xs) / len(xs), 4) if xs else None

    stats = {
        "tickets_settled": len(settled_tickets),
        "ticket_wins": wins,
        "ticket_losses": len(settled_tickets) - wins,
        "win_rate": round(wins / len(settled_tickets), 4) if settled_tickets else None,
        "staked": staked,
        "pnl": pnl,
        "roi": round(pnl / staked, 4) if staked else None,
        "mean_clv_novig": avg(clv_vals),
        "decisions_settled": len(settled),
        "process_worked": len(worked),
        "process_missed": len(missed),
        "fairlines_n": len(fairlines),
        "mean_err_pts": avg([f["err_pts"] for f in fairlines]),
        "max_err_pts": max((f["err_pts"] for f in fairlines), default=None),
        "clean_mean_err": avg(clean),
        "flagged_mean_err": avg(flagged),
        "brier_ours": avg(briers_ours),
        "brier_close": avg(briers_close),
    }

    findings_worked = [f"{d['event']}: {d['note']}" for d in worked]
    findings_missed = [f"{d['event']}: {d['note']}" for d in missed]
    if stats["clean_mean_err"] is not None and stats["flagged_mean_err"] is not None:
        findings_missed.append(
            "Calibration: flagged-source fair lines (conflict / date trap / "
            f"low-confidence close) missed the no-vig close by "
            f"{stats['flagged_mean_err']:.2f} pts on average vs "
            f"{stats['clean_mean_err']:.2f} pts for clean sources — source "
            "hygiene IS calibration."
        )

    return {
        "decisions": decisions,
        "fairlines": fairlines,
        "stats": stats,
        "findings": {"worked": findings_worked, "missed": findings_missed},
        "dates": sorted({s["date"] for s in record.get("slates", [])}),
    }


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------


def _header(ws, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(name=ARIAL, bold=True, size=10)
        cell.alignment = Alignment(horizontal="left")


def write_xlsx(record: dict[str, Any], path: str | Path) -> None:
    data = dataset(record)
    wb = Workbook()

    # --- Bets sheet: one row per engine decision (inputs, blue) ----------
    ws = wb.active
    ws.title = "Bets"
    _header(
        ws,
        1,
        [
            "Date", "Event", "Selection", "Type", "Target", "Fill", "Stake $",
            "Result", "P&L $", "CLV raw", "CLV no-vig", "Anchor", "Process",
            "Notes",
        ],
    )
    input_font = Font(name=ARIAL, size=10, color=BLUE)
    for r, d in enumerate(data["decisions"], start=2):
        values = [
            d["date"], d["event"], d["selection"], d["type"], d.get("target"),
            d.get("fill"), d.get("stake") or 0, d["result"], d.get("pnl") or 0,
            d.get("clv_raw"), d.get("clv_novig"), d.get("anchor", ""),
            d["process"], d.get("note", ""),
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = input_font
        ws.cell(row=r, column=5).number_format = "0.000"
        ws.cell(row=r, column=6).number_format = "0.000"
        ws.cell(row=r, column=7).number_format = "$#,##0"
        ws.cell(row=r, column=9).number_format = "$#,##0;($#,##0);-"
        ws.cell(row=r, column=10).number_format = "0.0%"
        ws.cell(row=r, column=11).number_format = "0.0%"
    for col, width in zip("ABCDEFGHIJKLMN", (11, 16, 26, 13, 7, 7, 10, 10, 10, 9, 10, 12, 9, 60)):
        ws.column_dimensions[col].width = width
    legend = ws.cell(row=len(data["decisions"]) + 3, column=1,
                     value="Blue cells = ledger data (regenerated by sharpods-tracker; "
                           "edit data/track_record.json, not this sheet). "
                           "Dashboard recomputes from these rows automatically. "
                           "To grade a bet you placed yourself: set Type=ticket, "
                           "Result=won or lost, and fill Stake $ and P&L $ — the "
                           "Dashboard picks it up. For an order that never filled, "
                           "set Result=no fill.")
    legend.font = Font(name=ARIAL, size=9, italic=True)
    legend.fill = PatternFill("solid", fgColor=YELLOW)

    # --- FairLines sheet: calibration rows with formula columns ----------
    ws = wb.create_sheet("FairLines")
    _header(
        ws,
        1,
        [
            "Date", "Event", "Our fair (home)", "Close no-vig (home)",
            "Error (pts)", "Home won (1/0)", "Brier ours", "Brier close",
            "Source flag",
        ],
    )
    for r, f in enumerate(data["fairlines"], start=2):
        rows = [
            f["date"], f["event"], f["our_fair"], f["close_novig"], None,
            1 if f["home_won"] else 0, None, None, f["flag"],
        ]
        for c, v in enumerate(rows, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = input_font if c in (1, 2, 3, 4, 6, 9) else Font(name=ARIAL, size=10)
        ws.cell(row=r, column=5, value=f"=ABS(C{r}-D{r})*100").font = Font(name=ARIAL, size=10)
        ws.cell(row=r, column=7, value=f"=(C{r}-F{r})^2").font = Font(name=ARIAL, size=10)
        ws.cell(row=r, column=8, value=f"=(D{r}-F{r})^2").font = Font(name=ARIAL, size=10)
        for col in (3, 4):
            ws.cell(row=r, column=col).number_format = "0.0000"
        ws.cell(row=r, column=5).number_format = "0.00"
        for col in (7, 8):
            ws.cell(row=r, column=col).number_format = "0.0000"
    for col, width in zip("ABCDEFGHI", (11, 26, 14, 16, 11, 13, 11, 11, 15)):
        ws.column_dimensions[col].width = width

    # --- Dashboard sheet: all formulas -----------------------------------
    ws = wb.create_sheet("Dashboard")
    title = ws.cell(row=1, column=1, value="SharpOds Tracker — Dashboard")
    title.font = Font(name=ARIAL, bold=True, size=14)
    note = ws.cell(row=2, column=1,
                   value="Every value below is a formula over Bets/FairLines: "
                         "add or settle rows there and this sheet recomputes.")
    note.font = Font(name=ARIAL, size=9, italic=True)

    rows: list[tuple[str, str, str]] = [
        ("Tickets settled", '=COUNTIFS(Bets!D:D,"ticket",Bets!H:H,"<>pending")', "0"),
        ("Ticket wins", '=COUNTIFS(Bets!D:D,"ticket",Bets!H:H,"won")', "0"),
        ("Win rate", "=IFERROR(B6/B5,0)", "0.0%"),
        ("Total staked", '=SUMIFS(Bets!G:G,Bets!D:D,"ticket",Bets!H:H,"<>pending")', "$#,##0"),
        ("Net P&L", "=SUM(Bets!I:I)", "$#,##0;($#,##0);-"),
        ("ROI on stakes", "=IFERROR(B9/B8,0)", "0.0%"),
        ("Mean CLV (no-vig), tickets", '=IFERROR(AVERAGEIFS(Bets!K:K,Bets!D:D,"ticket"),0)', "0.0%"),
        ("Limit orders unfilled", '=COUNTIF(Bets!D:D,"order_unfilled")', "0"),
        ("Open orders / pending", '=COUNTIF(Bets!H:H,"pending")', "0"),
        ("Decisions settled", '=COUNTIF(Bets!M:M,"worked")+COUNTIF(Bets!M:M,"miss")', "0"),
        ("Process validated (worked)", '=COUNTIF(Bets!M:M,"worked")', "0"),
        ("Process misses", '=COUNTIF(Bets!M:M,"miss")', "0"),
        ("Fair lines graded", "=COUNT(FairLines!E:E)", "0"),
        ("Mean |fair − close| (pts)", "=IFERROR(AVERAGE(FairLines!E:E),0)", "0.00"),
        ("Max |fair − close| (pts)", "=IFERROR(MAX(FairLines!E:E),0)", "0.00"),
        ("Clean-source mean error (pts)", '=IFERROR(AVERAGEIFS(FairLines!E:E,FairLines!I:I,"clean"),0)', "0.00"),
        ("Flagged-source mean error (pts)", '=IFERROR(AVERAGEIFS(FairLines!E:E,FairLines!I:I,"<>clean"),0)', "0.00"),
        ("Brier — SharpOds fair", "=IFERROR(AVERAGE(FairLines!G:G),0)", "0.0000"),
        ("Brier — closing line", "=IFERROR(AVERAGE(FairLines!H:H),0)", "0.0000"),
        ("Brier — coin flip (reference)", "=0.25", "0.0000"),
    ]
    start = 5
    for i, (label, formula, fmt) in enumerate(rows):
        r = start + i
        lab = ws.cell(row=r, column=1, value=label)
        lab.font = Font(name=ARIAL, size=10)
        val = ws.cell(row=r, column=2, value=formula)
        val.font = Font(name=ARIAL, size=10, bold=True, color=BLACK)
        val.number_format = fmt
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.cell(row=start, column=2).comment = Comment(
        "Win rate and P&L are outcome metrics; mean no-vig CLV is the process "
        "metric the model answers to (Squares & Sharps; Sharper). Small n: "
        "treat every number as provisional until the ledger's significance "
        "tests clear (see sharpods/clv.py).",
        "SharpOds",
    )

    r = start + len(rows) + 2
    head = ws.cell(row=r, column=1, value="What worked / what didn't (from the decision log)")
    head.font = Font(name=ARIAL, bold=True, size=11)
    data_findings = data["findings"]
    line = r + 1
    for kind, items in (("WORKED", data_findings["worked"]), ("DIDN'T", data_findings["missed"])):
        for item in items:
            c = ws.cell(row=line, column=1, value=f"[{kind}] {item}")
            c.font = Font(name=ARIAL, size=9)
            c.alignment = Alignment(wrap_text=True)
            line += 1
    src = ws.cell(row=line + 1, column=1,
                  value="Source: data/track_record.json (SharpOds ledger); regenerated by "
                        "`sharpods-tracker`. Assumption: closing lines are best-available "
                        "day-of published numbers, not exchange-verified closing ticks.")
    src.font = Font(name=ARIAL, size=8, italic=True)

    wb.save(path)


# ---------------------------------------------------------------------------
# Dashboard HTML
# ---------------------------------------------------------------------------


def render_dashboard(record: dict[str, Any], xlsx_path: str | Path | None = None) -> str:
    data = dataset(record)
    template = (
        resources.files("sharpods") / "templates" / "tracker.html.tmpl"
    ).read_text()
    csv_lines = ["date,event,selection,type,target,fill,stake,result,pnl,clv_raw,clv_novig,anchor,process"]
    for d in data["decisions"]:
        csv_lines.append(
            ",".join(
                "" if v is None else str(v).replace(",", ";")
                for v in (
                    d["date"], d["event"], d["selection"], d["type"],
                    d.get("target"), d.get("fill"), d.get("stake"), d["result"],
                    d.get("pnl"), d.get("clv_raw"), d.get("clv_novig"),
                    d.get("anchor"), d["process"],
                )
            )
        )
    payload = {**data, "csv": "\n".join(csv_lines)}
    return template.replace("__DATA_JSON__", json.dumps(payload))


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="sharpods-tracker")
    parser.add_argument("--record", default="data/track_record.json")
    parser.add_argument("--xlsx", default="data/tracker.xlsx")
    parser.add_argument("--html", default=None, help="also render the dashboard page here")
    parser.add_argument("--skip-xlsx", action="store_true")
    args = parser.parse_args(argv)

    record = load_record(args.record)
    if not args.skip_xlsx:
        write_xlsx(record, args.xlsx)
        print(f"wrote {args.xlsx}")
    if args.html:
        Path(args.html).write_text(render_dashboard(record, args.xlsx))
        print(f"wrote {args.html}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
