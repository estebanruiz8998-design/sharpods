from pathlib import Path

import pytest

from sharpods import tracker
from sharpods.ledger import load_record

RECORD_PATH = Path(__file__).resolve().parent.parent / "data" / "track_record.json"


@pytest.fixture(scope="module")
def record():
    return load_record(RECORD_PATH)


@pytest.fixture(scope="module")
def data(record):
    return tracker.dataset(record)


class TestDataset:
    def test_decisions_flattened_with_dates(self, data):
        assert len(data["decisions"]) >= 8
        assert all("date" in d for d in data["decisions"])

    def test_ticket_stats_match_ledger(self, data):
        s = data["stats"]
        assert s["tickets_settled"] == 1
        assert s["ticket_wins"] == 1
        assert s["win_rate"] == pytest.approx(1.0)
        assert s["staked"] == pytest.approx(30000.0)
        assert s["pnl"] == pytest.approx(29700.0)
        assert s["roi"] == pytest.approx(0.99)
        assert s["mean_clv_novig"] == pytest.approx(-0.0513)

    def test_calibration_split_shows_source_lesson(self, data):
        s = data["stats"]
        assert s["fairlines_n"] == 8
        # Clean sources track the close far better than flagged ones.
        assert s["clean_mean_err"] < 1.0
        assert s["flagged_mean_err"] > 4.0

    def test_findings_include_calibration_lesson(self, data):
        assert any("source" in f.lower() for f in data["findings"]["missed"])
        assert data["findings"]["worked"]

    def test_pending_decisions_not_counted_as_settled(self, data):
        s = data["stats"]
        settled_ids = [
            d["id"] for d in data["decisions"] if d["result"] != "pending"
        ]
        assert s["decisions_settled"] == len(settled_ids)


class TestWorkbook:
    def test_writes_three_sheets_with_formulas(self, record, tmp_path):
        from openpyxl import load_workbook

        path = tmp_path / "tracker.xlsx"
        tracker.write_xlsx(record, path)
        wb = load_workbook(path)
        assert set(wb.sheetnames) == {"Bets", "FairLines", "Dashboard"}
        # FairLines error column is a live formula, not a hardcoded value.
        assert wb["FairLines"]["E2"].value == "=ABS(C2-D2)*100"
        # Dashboard win rate divides the two COUNTIFS above it.
        b_col = [wb["Dashboard"].cell(row=r, column=2).value for r in range(5, 25)]
        assert any(isinstance(v, str) and v.startswith("=COUNTIFS(Bets!") for v in b_col)
        assert "=IFERROR(B6/B5,0)" in b_col


class TestDashboard:
    def test_renders_with_data_injected(self, record):
        html = tracker.render_dashboard(record)
        assert "__DATA_JSON__" not in html
        assert "SharpOds" in html
        assert "Cubs ML" in html
        assert "csv" in html

    def test_cli_smoke(self, tmp_path):
        xlsx = tmp_path / "t.xlsx"
        html = tmp_path / "t.html"
        assert tracker.main(["--record", str(RECORD_PATH), "--xlsx", str(xlsx), "--html", str(html)]) == 0
        assert xlsx.exists() and html.exists()
